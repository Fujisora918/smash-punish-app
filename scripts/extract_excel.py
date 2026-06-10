"""
Excel フレームデータ → JSON 変換スクリプト
Usage: python3 scripts/extract_excel.py
"""
import json
import openpyxl

EXCEL_PATH = "/Users/fujimurashouta/Downloads/スマブラSP データいろいろ (Ver.11.0.0).xlsx"
OUTPUT_PATH = "public/data/frame_data.json"

def safe_int(val):
    if val is None or val == '-':
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

def extract_fighters(wb):
    ws = wb["ガード硬直差(地上攻撃)"]
    fighters = []
    idx = 1
    for row in ws.iter_rows(min_row=3, values_only=True):
        name = row[1]
        if name is None or str(name) == "DLC11":
            continue
        fighters.append({
            "id": str(idx),
            "name_jp": str(name),
            "name_en": "",
            "file_id": "",
        })
        idx += 1
    return fighters

def extract_ground_moves(wb, fighters_map):
    ws = wb["ガード硬直差(地上攻撃)"]

    # 技名とカラムインデックスのマッピング (0-indexed)
    # 各技は4列: 発生(0), 全体(1), ダメ(2), 硬直差(3)
    MOVE_COLS = [
        ("弱1",       2),
        ("弱2",       6),
        ("弱3",       10),
        ("百裂〆",    14),
        ("DA",        18),
        ("横強",      22),
        ("上強",      26),
        ("下強",      30),
        ("横スマ",    34),
        ("上スマ",    38),
        ("下スマ",    42),
    ]

    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        no = row[0]
        name = row[1]
        if no is None or name is None:
            continue
        key = str(name)
        moves = {}
        for move_name, start_col in MOVE_COLS:
            startup = safe_int(row[start_col])
            total   = safe_int(row[start_col + 1])
            damage  = safe_int(row[start_col + 2])
            shield_adv = safe_int(row[start_col + 3])
            if startup is not None or shield_adv is not None:
                moves[move_name] = {
                    "startup": startup,
                    "total": total,
                    "damage": damage,
                    "shield_adv": shield_adv,
                }
        data[key] = moves
    return data

def extract_air_moves(wb):
    ws = wb["ガード硬直差(空中攻撃)"]

    # 各技は7列: 発生, 全体, ダメ, 着地隙, 着地隙発生, 硬直差(上り), 硬直差(最低空)
    MOVE_COLS = [
        ("空N",  2),
        ("空前", 9),
        ("空後", 16),
        ("空上", 23),
        ("空下", 30),
    ]

    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        no = row[0]
        name = row[1]
        if no is None or name is None:
            continue
        key = str(name)
        moves = {}
        for move_name, start_col in MOVE_COLS:
            startup       = safe_int(row[start_col])
            total         = safe_int(row[start_col + 1])
            damage        = safe_int(row[start_col + 2])
            landing_lag   = safe_int(row[start_col + 3])
            land_startup  = safe_int(row[start_col + 4])
            shield_rising = safe_int(row[start_col + 5])
            shield_low    = safe_int(row[start_col + 6])
            moves[move_name] = {
                "startup": startup,
                "total": total,
                "damage": damage,
                "landing_lag": landing_lag,
                "landing_startup": land_startup,
                "shield_adv_rising": shield_rising,
                "shield_adv_low": shield_low,
            }
        data[key] = moves
    return data

def extract_punish_moves(wb, sheet_name):
    ws = wb[sheet_name]

    # col 13-27 (0-indexed: 12-26) が各技の発生フレーム
    MOVE_NAMES = ["空N", "空前", "空後", "空上", "空下", "つかみ", "上スマ", "上B",
                  "弱", "DA", "横強", "上強", "下強", "横スマ", "下スマ"]

    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        no = row[0]
        name = row[1]
        if no is None or name is None:
            continue
        key = str(name)
        moves = {}
        for i, move_name in enumerate(MOVE_NAMES):
            val = safe_int(row[12 + i])
            moves[move_name] = val  # None = その技なし or "-"
        data[key] = moves
    return data

def main():
    print("Excelを読み込んでいます...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    fighters = extract_fighters(wb)
    print(f"  ファイター数: {len(fighters)}")

    fighters_map = {f["name_jp"]: f for f in fighters}

    ground_moves = extract_ground_moves(wb, fighters_map)
    print(f"  地上攻撃データ: {len(ground_moves)} キャラ")

    air_moves = extract_air_moves(wb)
    print(f"  空中攻撃データ: {len(air_moves)} キャラ")

    punish_front = extract_punish_moves(wb, "ガード硬直後最速行動(前)")
    punish_back  = extract_punish_moves(wb, "ガード硬直後最速行動(後)")
    print(f"  確反技データ(前): {len(punish_front)} キャラ")
    print(f"  確反技データ(後): {len(punish_back)} キャラ")

    output = {
        "fighters": fighters,
        "ground_moves": ground_moves,
        "air_moves": air_moves,
        "punish_moves_front": punish_front,
        "punish_moves_back": punish_back,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 出力完了: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
